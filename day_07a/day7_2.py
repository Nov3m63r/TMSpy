import json
import openpyxl

xls_file = 'task2_original.xlsx'
json_file = 'json_file7_2.json'

# data = {
#     'name': 'Alex',
#     'age': 24,
#     'marks': [7, 10, 9, 8, {'a': [1, 2, 3, [1, 2, [3]]]}],
#     'None': None,
# }

wb = openpyxl.load_workbook(xls_file)
sheet = wb.active
#print(sheet)
i = 1
fio = []
for row in sheet.iter_rows(min_row=1, max_col=2):
    # print(sheet[f'A{i}'])
    # fio = sheet[f'A{i}'].value
    fio.append(sheet[f'A{i}'].value + ' ' + sheet[f'B{i}'].value)
    i += 1
    print(fio)

# for row in sheet.iter_rows(min_row=1, max_col=2):
#     sheet.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)

# sheet.insert_cols(1)

# https://openpyxl.readthedocs.io/en/stable/tutorial.html
# +построчно считать лист
# +for row in ...
# +для каждой строки объединить 1 и 2 ячейку, сложить содержимое через пробел (Имя + Фамилия)
# +    сложить содержимое
#     объединить ячейки
#     присвоить результат объединенной ячейке
# перед первой колонкой вставить колонку с формулой - средний балл для каждого студента
# построчно считать оценки до слова end, вычислить средний балл
# в JSON файл сохранить словарь для всех студентов из таблицы:
# {
# 	“students”: {
# 		“Имя Фамилия”: {
# 			“marks”: [массив оценок],
# 			“average_mark”: средний_балл,
# 		},
# 		…
# 	},
# 	“total_average_mark”: средний_балл_между_всеми_студентами
# }
wb.save(xls_file)

# with open(json_file, 'w') as json_file:
#     json_str = json.dumps(data, indent=4)
#     print(json_str)
#     json_file.write(json_str)

# print("json file is ready")